"""
Sprint 3
Day 15

Financial Screener Engine

This module:

1. Reads screener_config.yaml
2. Connects to SQLite
3. Loads financial data
4. Applies configurable filters
5. Returns screened companies
6. Exports results
"""

import sqlite3
import yaml
import pandas as pd

from pathlib import Path
import numpy as np

from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Alignment

PROJECT_ROOT = Path(__file__).resolve().parents[2]

DATABASE_PATH = PROJECT_ROOT / "db" / "nifty100.db"

CONFIG_PATH = PROJECT_ROOT / "config" / "screener_config.yaml"

OUTPUT_PATH = PROJECT_ROOT / "output"

OUTPUT_PATH.mkdir(exist_ok=True)

def connect_db():
    """
    Create SQLite connection.
    """

    return sqlite3.connect(DATABASE_PATH)

def load_config():
    """
    Load screener configuration.
    """

    with open(CONFIG_PATH, "r") as file:
        config = yaml.safe_load(file)

    return config

def apply_single_filter(df, column, operator, value):
    """
    Apply a single filter only if the column exists.
    Returns the filtered DataFrame.
    """

    if column not in df.columns:
        print(f"[SKIPPED] Column '{column}' not available.")
        return df

    before = len(df)

    if operator == ">=":
        df = df[df[column] >= value]

    elif operator == "<=":
        df = df[df[column] <= value]

    elif operator == ">":
        df = df[df[column] > value]

    elif operator == "<":
        df = df[df[column] < value]

    elif operator == "==":
        df = df[df[column] == value]

    after = len(df)

    print(f"{column} {operator} {value} : {before} -> {after}")

    return df

def run_preset(df, preset_name, rules):
    """
    Apply one preset and return a filtered DataFrame.
    """

    print("\n")
    print("=" * 60)
    print(f"Running Preset : {preset_name}")
    print("=" * 60)

    result = df.copy()

    if "roe_min" in rules:
        result = apply_single_filter(
            result,
            "return_on_equity_pct",
            ">=",
            rules["roe_min"]
        )

    if "debt_to_equity_max" in rules:

        non_financial = result[
            result["broad_sector"] != "Financials"
        ]

        financial = result[
            result["broad_sector"] == "Financials"
        ]

        non_financial = apply_single_filter(
            non_financial,
            "debt_to_equity",
            "<=",
            rules["debt_to_equity_max"]
        )

        result = pd.concat(
            [non_financial, financial],
            ignore_index=True
        )

    if "free_cash_flow_min" in rules:
        result = apply_single_filter(
            result,
            "free_cash_flow_cr",
            ">=",
            rules["free_cash_flow_min"]
        )

    if "revenue_cagr_5yr_min" in rules:
        result = apply_single_filter(
            result,
            "revenue_cagr_5yr",
            ">=",
            rules["revenue_cagr_5yr_min"]
        )

    if "pat_cagr_5yr_min" in rules:
        result = apply_single_filter(
            result,
            "pat_cagr_5yr",
            ">=",
            rules["pat_cagr_5yr_min"]
        )

    if "dividend_payout_ratio_pct_max" in rules:
        result = apply_single_filter(
            result,
            "dividend_payout_ratio_pct",
            "<=",
            rules["dividend_payout_ratio_pct_max"]
        )

    if "debt_to_equity_exact" in rules:
        result = apply_single_filter(
            result,
            "debt_to_equity",
            "==",
            rules["debt_to_equity_exact"]
        )

    if "composite_quality_score" in result.columns:
        result = result.sort_values(
            "composite_quality_score",
            ascending=False
        )

    print(f"Companies Selected : {len(result)}")

    return result

def load_tables(conn):
    """
    Load all required tables.
    """

    financial_ratios = pd.read_sql_query(
        "SELECT * FROM financial_ratios",
        conn
    )

    companies = pd.read_sql_query(
        "SELECT * FROM companies",
        conn
    )

    sectors = pd.read_sql_query(
        "SELECT * FROM sectors",
        conn
    )

    stock_prices = pd.read_sql_query(
        "SELECT * FROM stock_prices",
        conn
    )

    print("\nTables Loaded")
    print("-" * 30)

    print(
        "financial_ratios :",
        len(financial_ratios)
    )

    print(
        "companies :",
        len(companies)
    )

    print(
        "sectors :",
        len(sectors)
    )

    print(
        "stock_prices :",
        len(stock_prices)
    )

    return (
        financial_ratios,
        companies,
        sectors,
        stock_prices
    )

def prepare_master_dataframe():
    """
    Prepare the master DataFrame used by the Screener Engine.

    Steps:
    1. Connect to SQLite
    2. Load required tables
    3. Get latest stock price for each company
    4. Merge financial ratios, companies, sectors and prices
    5. Keep only the latest financial year per company
    6. Perform validation checks
    7. Return the final master DataFrame
    """

    conn = connect_db()

    (
        financial_ratios,
        companies,
        sectors,
        stock_prices
    ) = load_tables(conn)

    # -------------------------------------------------------
    # Latest available stock price for each company
    # -------------------------------------------------------

    latest_prices = (
        stock_prices
        .sort_values("date")
        .groupby("company_id", as_index=False)
        .tail(1)
    )

    # -------------------------------------------------------
    # Merge financial ratios with company master
    # -------------------------------------------------------

    master = financial_ratios.merge(
        companies,
        left_on="company_id",
        right_on="id",
        how="left",
        suffixes=("", "_company")
    )
    # -------------------------------------------------------
# Remove records that do not match a valid company
# -------------------------------------------------------

    master = master[
    master["company_name"].notna()
].copy()

    # -------------------------------------------------------
    # Merge sector information
    # -------------------------------------------------------

    master = master.merge(
        sectors,
        on="company_id",
        how="left"
    )
    # -------------------------------------------------------
# Keep latest financial year only
# -------------------------------------------------------

    master = (
    master
    .sort_values("year")
    .groupby("company_id", as_index=False)
    .tail(1)
    .reset_index(drop=True)
)

    # -------------------------------------------------------
    # Merge latest stock prices
    # -------------------------------------------------------

    master = master.merge(
        latest_prices,
        on="company_id",
        how="left",
        suffixes=("", "_price")
    )

    # -------------------------------------------------------
    # Sort records
    # -------------------------------------------------------

    master = (
        master
        .sort_values(
            ["company_id", "year"]
        )
        .reset_index(drop=True)
    )

    # -------------------------------------------------------
    # Keep latest financial year only
    # -------------------------------------------------------

    master = (
        master
        .sort_values("year")
        .groupby("company_id", as_index=False)
        .tail(1)
        .reset_index(drop=True)
    )

    # -------------------------------------------------------
    # Validation Summary
    # -------------------------------------------------------

    print()
    print("=" * 60)
    print("MASTER DATASET")
    print("=" * 60)

    print(f"Latest Company Records : {len(master)}")
    print(f"Columns               : {len(master.columns)}")

    print()

    print("Unique Company IDs")
    print("-" * 40)

    print(
        "financial_ratios :",
        financial_ratios["company_id"].nunique()
    )

    if "company_id" in companies.columns:
        print(
            "companies        :",
            companies["company_id"].nunique()
        )
    else:
        print(
            "companies        :",
            companies["id"].nunique()
        )

    print(
        "master           :",
        master["company_id"].nunique()
    )

    print()

    print("Duplicate Company IDs")
    print("-" * 40)

    duplicates = (
        master["company_id"]
        .value_counts()
    )

    duplicates = duplicates[
        duplicates > 1
    ]

    if duplicates.empty:
        print("No duplicate company IDs found.")
    else:
        print(duplicates)

    print()

    print("Available Columns")
    print("-" * 60)

    for column in sorted(master.columns):
        print(column)

    conn.close()

    return master

    preset_results = {}
    presets = config["presets"]

    for preset_name, preset_rules in presets.items():
     preset_df = run_preset(
        master_df,
        preset_name,
        preset_rules
    )

    preset_results[preset_name] = preset_df
   

def apply_min_filter(df, column, minimum):
    """
    Keep rows where column >= minimum.
    Ignore filter if minimum is None.
    """

    if minimum is None:
        return df

    if column not in df.columns:
        print(f"[WARNING] Column '{column}' not found. Skipping filter.")
        return df

    before = len(df)

    df = df[df[column] >= minimum]

    after = len(df)

    print(f"{column} >= {minimum} : {before} -> {after}")

    return df

def apply_max_filter(df, column, maximum):
    """
    Keep rows where column <= maximum.
    """

    if maximum is None:
        return df

    if column not in df.columns:
        print(f"[WARNING] Column '{column}' not found. Skipping filter.")
        return df

    before = len(df)

    df = df[df[column] <= maximum]

    after = len(df)

    print(f"{column} <= {maximum} : {before} -> {after}")

    return df

def apply_debt_to_equity_filter(df, maximum):
    """
    Apply D/E filter.

    Financial companies always pass.
    """

    if maximum is None:
        return df

    before = len(df)

    financial_mask = (
        df["broad_sector"]
        == "Financials"
    )

    non_financial_mask = (
        ~financial_mask
    )

    filtered_non_financial = df[
        non_financial_mask
    ]

    filtered_non_financial = filtered_non_financial[
        filtered_non_financial["debt_to_equity"] <= maximum
    ]

    financial_df = df[
        financial_mask
    ]

    result = pd.concat(
        [
            financial_df,
            filtered_non_financial
        ]
    )

    after = len(result)

    print(f"D/E Filter : {before} -> {after}")

    return result

def apply_interest_coverage_filter(df, minimum):
    """
    Apply Interest Coverage filter.

    If icr_label exists:
        Debt Free companies always pass.

    Otherwise:
        Companies with NULL interest_coverage
        are treated as Debt Free.
    """

    if minimum is None:
        return df

    before = len(df)

    # Case 1: Sprint document implementation
    if "icr_label" in df.columns:

        debt_free = (
            df["icr_label"] == "Debt Free"
        )

    # Case 2: Your current database
    else:

        debt_free = (
            df["interest_coverage"].isna()
        )

    normal = ~debt_free

    filtered = df[normal]

    filtered = filtered[
        filtered["interest_coverage"] >= minimum
    ]

    result = pd.concat(
        [
            filtered,
            df[debt_free]
        ]
    )

    after = len(result)

    print(
        f"Interest Coverage Filter : {before} -> {after}"
    )

    return result

def apply_filters(df, config):
    """
    Apply all configured filters.
    """

    filters = config["filters"]

    # -------------------------------------------------
    # ROE
    # -------------------------------------------------

    df = apply_min_filter(
        df,
        "return_on_equity_pct",
        filters.get("roe_min")
    )

    # -------------------------------------------------
    # Debt to Equity
    # -------------------------------------------------

    df = apply_debt_to_equity_filter(
        df,
        filters.get("debt_to_equity_max")
    )

    # -------------------------------------------------
    # Free Cash Flow
    # -------------------------------------------------

    df = apply_min_filter(
        df,
        "free_cash_flow_cr",
        filters.get("free_cash_flow_min")
    )

    # -------------------------------------------------
    # Revenue CAGR
    # -------------------------------------------------

    df = apply_min_filter(
        df,
        "revenue_cagr_5yr",
        filters.get("revenue_cagr_5yr_min")
    )

    # -------------------------------------------------
    # PAT CAGR
    # -------------------------------------------------

    df = apply_min_filter(
        df,
        "pat_cagr_5yr",
        filters.get("pat_cagr_5yr_min")
    )

    # -------------------------------------------------
    # Operating Margin
    # -------------------------------------------------

    df = apply_min_filter(
        df,
        "operating_profit_margin_pct",
        filters.get("operating_profit_margin_min")
    )

    # -------------------------------------------------
    # Interest Coverage
    # -------------------------------------------------

    df = apply_interest_coverage_filter(
        df,
        filters.get("interest_coverage_min")
    )

   # -------------------------------------------------
# Metrics not available in current database
#
# These filters are intentionally skipped.
#
# They will be enabled when the database contains:
#
# - pe_ratio
# - pb_ratio
# - dividend_yield
# - market_cap
# - net_profit
# - sales
# -------------------------------------------------
    # -------------------------------------------------
    # EPS CAGR
    # -------------------------------------------------

    df = apply_min_filter(
        df,
        "eps_cagr_5yr",
        filters.get("eps_cagr_5yr_min")
    )

    # -------------------------------------------------
    # Asset Turnover
    # -------------------------------------------------

    df = apply_min_filter(
        df,
        "asset_turnover",
        filters.get("asset_turnover_min")
    )

    # -------------------------------------------------
    # Sales
    # -------------------------------------------------

    df = apply_min_filter(
        df,
        "sales",
        filters.get("sales_min")
    )

    return df

def sort_results(df, config):

    sorting = config["sorting"]

    return df.sort_values(

        sorting["column"],

        ascending=sorting["ascending"]

    )

def format_excel_sheet(ws):
    """
    Apply formatting to one worksheet.
    """

    green_fill = PatternFill(
        fill_type="solid",
        start_color="C6EFCE",
        end_color="C6EFCE"
    )

    red_fill = PatternFill(
        fill_type="solid",
        start_color="FFC7CE",
        end_color="FFC7CE"
    )

    header_fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78",
        end_color="1F4E78"
    )

    # -------------------------
    # Header Formatting
    # -------------------------

    for cell in ws[1]:

        cell.fill = header_fill

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    # -------------------------
    # Auto Width
    # -------------------------

    for column_cells in ws.columns:

        length = max(
            len(str(cell.value))
            if cell.value is not None
            else 0
            for cell in column_cells
        )

        ws.column_dimensions[
            column_cells[0].column_letter
        ].width = min(length + 3, 40)

    # -------------------------
    # Green / Red Formatting
    # -------------------------

    header_map = {
        cell.value: idx + 1
        for idx, cell in enumerate(ws[1])
    }

    green_rules = {
        "return_on_equity_pct": 15,
        "free_cash_flow_cr": 0,
        "revenue_cagr_5yr": 10,
        "pat_cagr_5yr": 10
    }

    red_rules = {
        "debt_to_equity": 2
    }

    for row in range(2, ws.max_row + 1):

        for column, limit in green_rules.items():

         if column in header_map:

          cell = ws.cell(
            row=row,
            column=header_map[column]
        )

        if isinstance(cell.value, (int, float)):

            if cell.value >= limit:

                cell.fill = green_fill

            else:

                cell.fill = red_fill
        for column, limit in red_rules.items():

            if column in header_map:

                cell = ws.cell(
                    row=row,
                    column=header_map[column]
                )

                if (
                    isinstance(cell.value, (int, float))
                    and cell.value <= limit
                 ):
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill

def winsorize_series(series):
    """
    Caps values between the 10th and 90th percentile.
    """

    lower = series.quantile(0.10)
    upper = series.quantile(0.90)

    return series.clip(lower=lower, upper=upper)

def normalize_score(series):
    """
    Converts a numeric series into a 0–100 score.
    """

    minimum = series.min()
    maximum = series.max()

    if maximum == minimum:
        return pd.Series(50, index=series.index)

    return ((series - minimum) / (maximum - minimum)) * 100

def calculate_composite_score(df):
    """
    Calculate a weighted composite quality score (0–100).
    """

    df = df.copy()

    metrics = [
        "return_on_equity_pct",
        "net_profit_margin_pct",
        "revenue_cagr_5yr",
        "pat_cagr_5yr",
        "asset_turnover",
        "interest_coverage",
        "debt_to_equity",
        "free_cash_flow_cr"
    ]

    for metric in metrics:

        if metric not in df.columns:
            continue

        df[metric] = winsorize_series(df[metric])

        if metric == "debt_to_equity":

            score = normalize_score(df[metric])

            df["score_" + metric] = 100 - score

        else:

            df["score_" + metric] = normalize_score(df[metric])

    df["composite_quality_score"] = (

        df.get("score_return_on_equity_pct", 0) * 0.15 +

        df.get("score_net_profit_margin_pct", 0) * 0.10 +

        df.get("score_free_cash_flow_cr", 0) * 0.20 +

        df.get("score_revenue_cagr_5yr", 0) * 0.10 +

        df.get("score_pat_cagr_5yr", 0) * 0.10 +

        df.get("score_asset_turnover", 0) * 0.10 +

        df.get("score_interest_coverage", 0) * 0.10 +

        df.get("score_debt_to_equity", 0) * 0.15
    )

    return df

def calculate_sector_scores(df):
    """
    Compute sector-relative composite score (0–100).
    """

    df = df.copy()

    sector_scores = []

    for sector, sector_df in df.groupby("broad_sector"):

        sector_df = sector_df.copy()

        minimum = sector_df["composite_quality_score"].min()

        maximum = sector_df["composite_quality_score"].max()

        if maximum == minimum:

            sector_df["sector_composite_score"] = 50

        else:

            sector_df["sector_composite_score"] = (

                (
                    sector_df["composite_quality_score"] - minimum
                )

                /

                (

                    maximum - minimum

                )

            ) * 100

        sector_scores.append(sector_df)

    return pd.concat(
        sector_scores,
        ignore_index=True
    )


def main():
 if __name__ == "__main__":

    print("\n" + "=" * 60)
    print("SPRINT 3 - DAY 15 & DAY 16 : FINANCIAL SCREENER ENGINE")
    print("=" * 60)

    # ==================================================
    # Load Configuration
    # ==================================================

    config = load_config()

    print("\nConfiguration loaded successfully.")

    print("\nCONFIG KEYS")
    print("-" * 30)
    print(config.keys())

    print("\nFULL CONFIG")
    print(config)

    # ==================================================
    # Prepare Master Data
    # ==================================================

    master_df = prepare_master_dataframe()

    print("Master DataFrame prepared successfully.")
    master_df = calculate_composite_score(master_df)

    print("\nComposite score calculated successfully.")

    master_df = calculate_sector_scores(master_df)

    print("Sector-relative score calculated successfully.")
    EXPORT_COLUMNS = [
    "company_id",
    "company_name",
    "broad_sector",
    "sub_sector",
    "year",

    "return_on_equity_pct",
    "roce_percentage",
    "net_profit_margin_pct",
    "operating_profit_margin_pct",

    "debt_to_equity",
    "interest_coverage",
    "asset_turnover",

    "free_cash_flow_cr",
    "capex_cr",
    "cash_from_operations_cr",

    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",

    "composite_quality_score",
    "sector_composite_score"
]

    print("\nTop Composite Scores")

    print(
    master_df[
        [
            "company_name",
            "composite_quality_score"
        ]
    ]
    .sort_values(
        "composite_quality_score",
        ascending=False
    )
    .head(10)
    )
    print(
    master_df[
        [
            "company_name",
            "broad_sector",
            "composite_quality_score",
            "sector_composite_score"
        ]
    ]
    .sort_values(
        [
            "broad_sector",
            "sector_composite_score"
        ],
        ascending=False
    )
    .head(20)

)

    # ==================================================
    # DAY 15 : CUSTOM SCREENER
    # ==================================================

    print("\n" + "=" * 60)
    print("DAY 15 : CUSTOM SCREENER")
    print("=" * 60)

    screened = apply_filters(
        master_df,
        config
    )

    print("\nFiltering completed successfully.")

    screened = sort_results(
        screened,
        config
    )

    print("\nSorting completed successfully.")

    print("\n" + "=" * 60)
    print("DAY 15 SUMMARY")
    print("=" * 60)

    print(f"Total Companies Analysed : {len(master_df)}")
    print(f"Companies Passed         : {len(screened)}")

    if len(master_df) > 0:
        selection_rate = (len(screened) / len(master_df)) * 100
        print(f"Selection Rate           : {selection_rate:.2f}%")

    print("\nTop Screened Companies")
    print("-" * 60)

    display_columns = [
        "company_id",
        "company_name",
        "year",
        "return_on_equity_pct",
        "debt_to_equity",
        "free_cash_flow_cr",
        "revenue_cagr_5yr",
        "pat_cagr_5yr",
        "operating_profit_margin_pct",
        "interest_coverage",
        "asset_turnover",
        "composite_quality_score"
    ]

    available_columns = [
        col for col in display_columns
        if col in screened.columns
    ]

    print(
        screened[available_columns]
        .head(10)
        .to_string(index=False)
    )

    # ==================================================
    # DAY 16 : PRESET SCREENERS
    # ==================================================

    print("\n" + "=" * 60)
    print("DAY 16 : PRESET SCREENERS")
    print("=" * 60)

    preset_results = {}

    for preset_name, preset_rules in config["presets"].items():

        preset_df = run_preset(
            master_df,
            preset_name,
            preset_rules
        )

        preset_df = sort_results(
            preset_df,
            config
        )

        preset_results[preset_name] = preset_df

    # ==================================================
    # EXPORT ALL PRESETS
    # ==================================================
    available_columns = [
    col
    for col in EXPORT_COLUMNS
    if col in preset_df.columns
]

    preset_df = preset_df[available_columns]
    output_file = OUTPUT_PATH / "screener_output.xlsx"

    validation_rows = []

    with pd.ExcelWriter(
        output_file,
        engine="openpyxl"
    ) as writer:

        for preset_name, preset_df in preset_results.items():

            sheet_name = (
                preset_name
                .replace("_", " ")
                .title()
                .replace(" ", "_")
            )[:31]
            available_columns = [
    col
    for col in EXPORT_COLUMNS
    if col in preset_df.columns
]

            preset_df = preset_df[available_columns]
            preset_df.to_excel(
    writer,
    sheet_name=sheet_name,
    index=False
)

            row_count = len(preset_df)

            status = (
                "PASS"
                if 5 <= row_count <= 50
                else "REVIEW"
            )

            validation_rows.append(
                {
                    "Preset": sheet_name,
                    "Companies": row_count,
                    "Status": status
                }
            )

        validation_df = pd.DataFrame(validation_rows)

        validation_df.to_excel(
            writer,
            sheet_name="Validation_Summary",
            index=False
        )

        workbook = writer.book

        for sheet in workbook.sheetnames:

            worksheet = workbook[sheet]

            format_excel_sheet(worksheet)

    # ==================================================
    # DAY 16 VALIDATION
    # ==================================================

    print("\n" + "=" * 60)
    print("DAY 16 VALIDATION")
    print("=" * 60)

    for row in validation_rows:

        print(
            f"{row['Preset']:<30}"
            f"{row['Companies']:>5} companies   "
            f"[{row['Status']}]"
        )

    # ==================================================
    # EXPORT SUMMARY
    # ==================================================

    print("\n" + "=" * 60)
    print("EXPORT COMPLETED")
    print("=" * 60)

    print(f"Workbook : {output_file}")
    print(f"Worksheets : {len(preset_results) + 1}")

    print("\nGenerated Sheets:")

    for row in validation_rows:

        print(
            f"  • {row['Preset']} ({row['Companies']} companies)"
        )

    print("  • Validation_Summary")

    print("\nWorkbook saved successfully.")

    # ==================================================
    # FINAL STATUS
    # ==================================================
    
    print("\n" + "=" * 60)
    print("SPRINT 3 STATUS")
    print("=" * 60)
    print("\nComposite Score Validation")
    print("-" * 40)
    print(f"Minimum Score : {master_df['composite_quality_score'].min():.2f}")
    print(f"Maximum Score : {master_df['composite_quality_score'].max():.2f}")
    print(f"Average Score : {master_df['composite_quality_score'].mean():.2f}")

    print("\nSector Relative Score Validation")
    print("-" * 40)

    sector_summary = (
     master_df.groupby("broad_sector")["sector_composite_score"]
    .agg(["min", "max", "mean"])
    .round(2)
)

    print(sector_summary)
    print("\nWorkbook Validation")
    print("-" * 40)
    print(f"Workbook Exists : {output_file.exists()}")
    print("\nOutput Generated:")
    print(output_file)

if __name__ == "__main__":
    main()
