"""
=========================================================
Sprint 3 - Day 18
Peer Percentile Ranking Engine
=========================================================

Loads peer groups from SQLite,
joins company financial data,
prepares peer comparison dataset,
creates peer_percentiles table.

Author : Bluestock Project
"""

import sqlite3
import pandas as pd
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os


# -------------------------------------------------------
# Project Paths
# -------------------------------------------------------

PROJECT_ROOT = Path(__file__).resolve().parents[2]

DATABASE_PATH = PROJECT_ROOT / "db" / "nifty100.db"


# -------------------------------------------------------
# SQLite Connection
# -------------------------------------------------------

def connect_database():
    """
    Connect to SQLite database.
    """

    print("\nConnecting to SQLite...")

    connection = sqlite3.connect(DATABASE_PATH)

    print("Connection Successful.")

    return connection


# -------------------------------------------------------
# Load Tables
# -------------------------------------------------------

def load_tables(connection):

    print("\nLoading Tables...")

    financial_ratios = pd.read_sql(
        "SELECT * FROM financial_ratios",
        connection
    )

    peer_groups = pd.read_sql(
        "SELECT * FROM peer_groups",
        connection
    )

    companies = pd.read_sql(
        "SELECT * FROM companies",
        connection
    )

    sectors = pd.read_sql(
        "SELECT * FROM sectors",
        connection
    )

    print("\nTables Loaded")

    print("----------------------------")
    print("financial_ratios :", len(financial_ratios))
    print("peer_groups      :", len(peer_groups))
    print("companies        :", len(companies))
    print("sectors          :", len(sectors))

    return (
        financial_ratios,
        peer_groups,
        companies,
        sectors,
    )


# -------------------------------------------------------
# Merge Tables
# -------------------------------------------------------

def build_master_dataframe(
    financial_ratios,
    peer_groups,
    companies,
    sectors,
):

    print("\nPreparing Master DataFrame...")

    df = financial_ratios.merge(
        peer_groups,
        on="company_id",
        how="left"
    )

    df = df.merge(
        companies[
            [
                "id",
                "company_name",
                "roce_percentage",
            ]
        ],
        left_on="company_id",
        right_on="id",
        how="left"
    )

    df = df.merge(
        sectors[
            [
                "company_id",
                "broad_sector",
                "sub_sector",
            ]
        ],
        on="company_id",
        how="left"
    )

    print("Rows :", len(df))
    print("Columns :", len(df.columns))

    return df


# -------------------------------------------------------
# Create Output Table
# -------------------------------------------------------

def create_peer_table(connection):

    cursor = connection.cursor()

    cursor.execute("""

    CREATE TABLE IF NOT EXISTS peer_percentiles(

        id INTEGER PRIMARY KEY AUTOINCREMENT,

        company_id TEXT,

        peer_group_name TEXT,

        metric TEXT,

        value REAL,

        percentile_rank REAL,

        year INTEGER

    );

    """)

    connection.commit()

    print("\npeer_percentiles table ready.")


# -------------------------------------------------------
# Show Missing Peer Groups
# -------------------------------------------------------

def show_missing_peer_groups(df):

    missing = df[
        df["peer_group_name"].isna()
    ]

    companies = missing["company_id"].nunique()

    print("\nCompanies Without Peer Group")

    print("----------------------------")

    if companies == 0:

        print("All companies assigned.")

    else:

        print(companies)

        print(
            "\nThese companies will be skipped "
            "during percentile calculation."
        )

# -------------------------------------------------------
# Metrics for Peer Ranking
# -------------------------------------------------------

RANKING_METRICS = {

    "return_on_equity_pct": False,

    "roce_percentage": False,

    "net_profit_margin_pct": False,

    "debt_to_equity": True,

    "free_cash_flow_cr": False,

    "pat_cagr_5yr": False,

    "revenue_cagr_5yr": False,

    "eps_cagr_5yr": False,

    "interest_coverage": False,

    "asset_turnover": False,

}

def calculate_percentiles(df):

    """
    Calculates percentile rankings
    for every peer group.
    """

    print("\nCalculating Peer Percentiles...")

    results = []

    grouped = df.groupby("peer_group_name")

    for peer_name, peer_df in grouped:

        if pd.isna(peer_name):

            continue

        print(f"\nPeer Group : {peer_name}")

        print("Companies :", peer_df["company_id"].nunique())

        for metric, inverse in RANKING_METRICS.items():

            if metric not in peer_df.columns:

                print(f"Skipping {metric}")

                continue

            temp = peer_df.copy()

            temp = temp[temp[metric].notna()]

            if len(temp) == 0:

                continue

            temp["percentile_rank"] = temp[metric].rank(
                pct=True,
                method="average"
            )

            if inverse:

                temp["percentile_rank"] = (
                    1 -
                    temp["percentile_rank"]
                )

            for _, row in temp.iterrows():

                results.append({

                    "company_id":
                        row["company_id"],

                    "peer_group_name":
                        peer_name,

                    "metric":
                        metric,

                    "value":
                        row[metric],

                    "percentile_rank":
                        round(
                            row["percentile_rank"],
                            4
                        ),

                    "year":
                        row["year"]

                })

    results = pd.DataFrame(results)

    print("\nTotal Rankings Created :", len(results))

    return results

def save_percentiles(connection, results):

    print("\nSaving Rankings...")

    connection.execute(
        "DELETE FROM peer_percentiles"
    )

    results.to_sql(

        "peer_percentiles",

        connection,

        if_exists="append",

        index=False

    )

    connection.commit()

    print("Database Updated Successfully.")

# -------------------------------------------------------
# Main
# -------------------------------------------------------

def main():

    print("=" * 60)
    print("SPRINT 3 - DAY 18")
    print("PEER PERCENTILE ENGINE")
    print("=" * 60)

    connection = connect_database()

    (
        financial_ratios,
        peer_groups,
        companies,
        sectors,
    ) = load_tables(connection)

    df = build_master_dataframe(
        financial_ratios,
        peer_groups,
        companies,
        sectors,
    )

    create_peer_table(connection)

    show_missing_peer_groups(df)
    rankings = calculate_percentiles(df)

    save_percentiles(
    connection,
    rankings
)

    print("\nRanking Preview\n")

    print(rankings.head(20))

    print("\nMaster Dataset Preview\n")

    print(
        df[
            [
                "company_id",
                "company_name",
                "peer_group_name",
                "return_on_equity_pct",
                "roce_percentage",
                "debt_to_equity",
            ]
        ].head(15)
    )
    # day-18 validation

    print("\n")
    print("=" * 60)
    print("DAY 18 VALIDATION")
    print("=" * 60)

    validation = pd.read_sql("""

SELECT

peer_group_name,

COUNT(*) total_records,

ROUND(AVG(percentile_rank),3) avg_percentile,

ROUND(MAX(percentile_rank),3) max_percentile,

ROUND(MIN(percentile_rank),3) min_percentile

FROM peer_percentiles

GROUP BY peer_group_name

ORDER BY peer_group_name

""", connection)

    print(validation)

    print("\n")

    print("Total Percentile Records :")

    count = pd.read_sql(

    "SELECT COUNT(*) total FROM peer_percentiles",

     connection

    )

    print(count.iloc[0,0])
    print("\n")
    print("=" * 60)
    print("CREATING PEER COMPARISON WORKBOOK")
    print("=" * 60)

    output_file = "output/peer_comparison.xlsx"

    os.makedirs("output", exist_ok=True)

    writer = pd.ExcelWriter(
    output_file,
    engine="openpyxl"
)

    peer_groups = pd.read_sql("""

SELECT DISTINCT peer_group_name
FROM peer_percentiles
ORDER BY peer_group_name

""", connection)

    for group in peer_groups["peer_group_name"]:

     df = pd.read_sql(
        """

        SELECT

        company_id,
        metric,
        value,
        percentile_rank,
        year

        FROM peer_percentiles

        WHERE peer_group_name=?

        ORDER BY
        company_id,
        metric,
        year

        """,
        connection,
        params=[group]
    )

    sheet = group[:31]

    df.to_excel(
        writer,
        sheet_name=sheet,
        index=False
    )

    writer.close()

    print("Workbook Created Successfully.")
    
    workbook = load_workbook(output_file)

    green = PatternFill(
    fill_type="solid",
    start_color="C6EFCE"
)

    yellow = PatternFill(
    fill_type="solid",
    start_color="FFF2CC"
)

    red = PatternFill(
    fill_type="solid",
    start_color="FFC7CE"
)

    for ws in workbook.worksheets: 
     headers = [c.value for c in ws[1]]

     if "percentile_rank" not in headers:
        continue

    col = headers.index("percentile_rank") + 1

    for row in range(2, ws.max_row + 1):

        value = ws.cell(row=row, column=col).value

        if value is None:
            continue

        if value >= 0.80:
            ws.cell(row=row, column=col).fill = green

        elif value >= 0.50:
            ws.cell(row=row, column=col).fill = yellow

        else:
            ws.cell(row=row, column=col).fill = red

    workbook.save(output_file)

    print("Colour Coding Completed.")
    connection.close()

    print("\nDay 18 Peer Percentile Ranking Completed Successfully")

    print("\n")
    print("=" * 60)
    print("PEER WORKBOOK SUMMARY")
    print("=" * 60)
    print("Workbook :", os.path.abspath(output_file))
    print("Sheets   :", len(workbook.sheetnames))

    for s in workbook.sheetnames:
     print(" •", s)


if __name__ == "__main__":
    main()
