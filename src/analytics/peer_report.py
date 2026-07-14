"""
============================================================
SPRINT 3 - DAY 20
Peer Comparison Excel Report
PART 1
------------------------------------------------------------
Loads all required tables
Creates latest company dataset
Merges peer groups
Prepares data for Excel generation
============================================================
"""

import sqlite3
from pathlib import Path

import pandas as pd
import os 
from openpyxl.styles import PatternFill, Font
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------
# PROJECT PATHS
# ----------------------------------------------------------

PROJECT_ROOT = Path(__file__).resolve().parents[2]

DATABASE = PROJECT_ROOT / "db" / "nifty100.db"

OUTPUT_FOLDER = PROJECT_ROOT / "output"
OUTPUT_FOLDER.mkdir(exist_ok=True)

OUTPUT_FILE = OUTPUT_FOLDER / "peer_comparison.xlsx"

print("\n============================================================")
print("SPRINT 3 - DAY 20")
print("============================================================")

# ----------------------------------------------------------
# CONNECT SQLITE
# ----------------------------------------------------------

connection = sqlite3.connect(DATABASE)

print("\nLoading SQLite Database...")

# ----------------------------------------------------------
# LOAD TABLES
# ----------------------------------------------------------

financial_ratios = pd.read_sql(
    "SELECT * FROM financial_ratios",
    connection
)

peer_groups = pd.read_sql(
    "SELECT * FROM peer_groups",
    connection
)

companies = pd.read_sql(
    "SELECT * FROM companies",
    connection
)

peer_percentiles = pd.read_sql(
    "SELECT * FROM peer_percentiles",
    connection
)

print("\nTables Loaded")
print("-------------------------------------")

print("financial_ratios :", len(financial_ratios))
print("peer_groups      :", len(peer_groups))
print("companies        :", len(companies))
print("peer_percentiles :", len(peer_percentiles))

# ----------------------------------------------------------
# LATEST FINANCIAL YEAR
# ----------------------------------------------------------

latest_year = financial_ratios["year"].max()

latest_ratios = financial_ratios[
    financial_ratios["year"] == latest_year
].copy()

print("\nLatest Financial Year :", latest_year)
print("Latest Records        :", len(latest_ratios))

# ----------------------------------------------------------
# MERGE PEER GROUPS
# ----------------------------------------------------------

master = latest_ratios.merge(
    peer_groups,
    on="company_id",
    how="left"
)

master = master.merge(
    companies[
        [
            "id",
            "company_name",
            "roce_percentage"
        ]
    ],
    left_on="company_id",
    right_on="id",
    how="left"
)

print("\nMaster Dataset Created")

print("-------------------------------------")

print("Rows    :", len(master))
print("Columns :", len(master.columns))

print("\nColumns")

for column in master.columns:
    print(column)

# ----------------------------------------------------------
# PEER GROUP SUMMARY
# ----------------------------------------------------------

peer_summary = (
    master.groupby("peer_group_name")
    .size()
    .reset_index(name="companies")
    .sort_values("peer_group_name")
)

print("\n============================================================")
print("PEER GROUP SUMMARY")
print("============================================================")

print(peer_summary)

print("\nPeer Groups Found :", len(peer_summary))

# ----------------------------------------------------------
# PART 1 VALIDATION
# ----------------------------------------------------------

print("\n============================================================")
print("DAY 20 - LOADED COMPLETED")
print("============================================================")

print("SQLite Connection     : OK")
print("Tables Loaded         : OK")
print("Latest Year Selected  : OK")
print("Master Dataset Ready  : OK")
print("Peer Groups Detected  :", len(peer_summary))

print("\nReady for Part 2 (Excel Workbook Generation)")


# ==========================================================
# DAY 20 
# CREATE EXCEL WORKBOOK
# ==========================================================

from openpyxl import Workbook

print("\n============================================================")
print("CREATING EXCEL WORKBOOK")
print("============================================================")

workbook = Workbook()

# Remove default worksheet
default_sheet = workbook.active
workbook.remove(default_sheet)

print("Workbook Created Successfully")

# ==========================================================
# METRICS REQUIRED IN REPORT
# ==========================================================

metrics = [

    "return_on_equity_pct",
    "roce_percentage",
    "net_profit_margin_pct",
    "debt_to_equity",
    "free_cash_flow_cr",
    "pat_cagr_5yr",
    "revenue_cagr_5yr",
    "eps_cagr_5yr",
    "interest_coverage",
    "asset_turnover",
    "composite_quality_score"

]

# ==========================================================
# CREATE ONE SHEET PER PEER GROUP
# ==========================================================

peer_groups_list = sorted(
    master["peer_group_name"]
    .dropna()
    .unique()
)

print("\nGenerating Worksheets...\n")

for group in peer_groups_list:

    print("Creating :", group)

    ws = workbook.create_sheet(title=group[:31])

    group_df = (
        master[
            master["peer_group_name"] == group
        ]
        .copy()
        .sort_values(
            "composite_quality_score",
            ascending=False
        )
    )

    headers = [

        "company_id",
        "company_name",
        "year"

    ]

    headers.extend(metrics)

    for metric in metrics:
        headers.append(metric + "_percentile")

    headers.append("Benchmark")

    ws.append(headers)

    # ------------------------------------------
    # Write company rows
    # ------------------------------------------

    for _, row in group_df.iterrows():

        excel_row = [

            row["company_id"],
            row["company_name"],
            row["year"]

        ]

        for metric in metrics:

            excel_row.append(
                row.get(metric)
            )

        # Percentiles
        for metric in metrics:

            value = peer_percentiles[
                (
                    peer_percentiles["company_id"] == row["company_id"]
                )
                &
                (
                    peer_percentiles["metric"] == metric
                )
                &
                (
                    peer_percentiles["year"] == row["year"]
                )
            ]

            if len(value):

                excel_row.append(
                    round(
                        value.iloc[0]["percentile_rank"],
                        4
                    )
                )

            else:

                excel_row.append(None)

        excel_row.append(row["is_benchmark"])

        ws.append(excel_row)

        # ==========================================================
# HEADER FORMATTING
# ==========================================================

    from openpyxl.styles import PatternFill, Font

    BLUE_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78"
)

    WHITE_FONT = Font(
    color="FFFFFF",
    bold=True
)

    GREEN_FILL = PatternFill(
    fill_type="solid",
    fgColor="C6EFCE"
)

    YELLOW_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFF2CC"
)

    RED_FILL = PatternFill(
    fill_type="solid",
    fgColor="F4CCCC"
)

    GOLD_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFD966"
)

# Header row

    for cell in ws[1]:
      cell.fill = BLUE_FILL
      cell.font = WHITE_FONT

    # ==========================================================
# PERCENTILE CELL COLOURING
# ==========================================================

    percentile_start = 3 + len(metrics) + 1
    percentile_end = percentile_start + len(metrics) - 1
    for r in range(2, ws.max_row + 1):
     for c in range(percentile_start, percentile_end + 1):
        value = ws.cell(r, c).value

        if value is None:
            continue

        if value >= 0.75:

            ws.cell(r, c).fill = GREEN_FILL

        elif value <= 0.25:

            ws.cell(r, c).fill = RED_FILL

        else:

            ws.cell(r, c).fill = YELLOW_FILL

            # ==========================================================
# HIGHLIGHT BENCHMARK COMPANY
# ==========================================================

    benchmark_column = ws.max_column

    for r in range(2, ws.max_row + 1):
     if ws.cell(r, benchmark_column).value == 1:

        for c in range(1, benchmark_column + 1):

            ws.cell(r, c).fill = GOLD_FILL

            # ==========================================================
# MEDIAN SUMMARY ROW
# ==========================================================

    median_row = []
    median_row.append("")
    median_row.append("Peer Group Median")
    median_row.append("")
    for metric in metrics:
     median_row.append(
        round(
            group_df[metric].median(),
            2
        )
    )

    for metric in metrics:

     median_row.append("")

     median_row.append("")

     ws.append(median_row)

    summary_row = ws.max_row

    for cell in ws[summary_row]:
     cell.fill = BLUE_FILL
     cell.font = WHITE_FONT

    # ==========================================================
# AUTO FIT COLUMN WIDTHS
# ==========================================================

    from openpyxl.utils import get_column_letter

    for sheet in workbook.worksheets:

     for column in sheet.columns:

         length = 0

         letter = get_column_letter(column[0].column)

         for cell in column:

            try:

                length = max(
                    length,
                    len(str(cell.value))
                )

            except:

                pass

         sheet.column_dimensions[
            letter
        ].width = min(length + 3, 35)

    print("\nAll Worksheets Created")
    print("Total Worksheets :", len(workbook.sheetnames)) 

# ==========================================================
# SAVE WORKBOOK
# ==========================================================
workbook.save(OUTPUT_FILE)

print("\nWorkbook Saved")

print(OUTPUT_FILE)
print("\nDAY 20 excel workbook COMPLETED")
